import os
import csv
import io
from datetime import datetime, timedelta

from flask import (
    Blueprint, render_template, redirect, url_for,
    flash, request, current_app, send_file, jsonify,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from models import (
    db, Ticket, User, Role, TicketComment, SLAPolicy,
    TicketAttachment, CSATRating, Tag, TicketWatcher, CannedResponse,
    TicketLink, CustomField, TicketFieldValue, TimeEntry, CallLog,
)
from utils import (
    send_notification, send_ticket_notification, allowed_file, scan_file_magic,
    create_notification, notify_staff, log_history,
    notify_slack, notify_teams,
)

tickets = Blueprint("tickets", __name__, url_prefix="/tickets")

STATUSES   = ["Open", "In Progress", "Pending", "Resolved", "Closed", "Reopened"]
PRIORITIES = ["Low", "Medium", "High", "Critical", "Urgent"]
PLATFORMS  = ["SPICE Web", "SPICE Mobile", "SPICE Admin", "CHW Tool", "Other"]
ISSUE_TYPES = [
    "Bug", "Feature Request", "Performance Issue", "Data Issue",
    "Login / Access Issue", "UI Issue", "Configuration Issue", "Other",
]
PRODUCTS = ["SPICE", "CHW Tool", "Admin Portal", "Other"]


# ── Create ─────────────────────────────────────────────────────────────────────

@tickets.route("/create", methods=["GET", "POST"])
@login_required
def create():
    from models import Role
    if current_user.role == Role.VIEWER:
        flash("Viewers cannot create tickets.", "warning")
        return redirect(url_for("dashboard.main"))
    if request.method == "POST":
        screenshot_path = None
        if "screenshot" in request.files:
            file = request.files["screenshot"]
            if file and file.filename and allowed_file(file.filename):
                safe, reason = scan_file_magic(file)
                if not safe:
                    flash(f"Screenshot rejected: {reason}", "danger")
                    return redirect(url_for("tickets.create"))
                filename = secure_filename(
                    f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                )
                upload_folder = current_app.config["UPLOAD_FOLDER"]
                os.makedirs(upload_folder, exist_ok=True)
                file.save(os.path.join(upload_folder, filename))
                screenshot_path = f"screenshots/{filename}"

        # Issue start date: auto-set to submission time (no longer collected from form)
        issue_start = datetime.utcnow()

        # Structured location
        country_id = request.form.get("country_id", type=int) or None
        admin1_id = request.form.get("admin1_id", type=int) or None
        admin2_id = request.form.get("admin2_id", type=int) or None
        admin3_id = request.form.get("admin3_id", type=int) or None

        # Build legacy district/upazila strings
        district_str = request.form.get("district", "").strip()
        upazila_str = request.form.get("upazila", "").strip()
        if not district_str and admin2_id:
            from models import AdminLevel2
            a2 = AdminLevel2.query.get(admin2_id)
            district_str = a2.name if a2 else ""
        if not upazila_str and admin3_id:
            from models import AdminLevel3
            a3 = AdminLevel3.query.get(admin3_id)
            upazila_str = a3.name if a3 else ""

        # Issue taxonomy
        category_id = request.form.get("category_id", type=int) or None
        subcategory_id = request.form.get("subcategory_id", type=int) or None
        # issue_type from subcategory name for legacy compatibility
        issue_type_str = request.form.get("issue_type", "")
        if subcategory_id:
            from models import IssueSubcategory
            sub = IssueSubcategory.query.get(subcategory_id)
            if sub:
                issue_type_str = sub.name

        ticket = Ticket(
            sl_no=Ticket.generate_sl_no(),
            channel="web",
            district=district_str,
            upazila=upazila_str,
            country_id=country_id,
            admin1_id=admin1_id,
            admin2_id=admin2_id,
            admin3_id=admin3_id,
            category_id=category_id,
            subcategory_id=subcategory_id,
            reporter_id=current_user.id,
            issue_reported_by_role=current_user.role.value,
            issue_reporter_name=current_user.full_name,
            issue_reporter_contact=(
                current_user.contact or request.form.get("contact", "").strip()
            ),
            issue_start_date=issue_start,
            spice_platform=request.form.get("spice_platform", ""),
            issue_type=issue_type_str,
            problem_details=request.form.get("problem_details", ""),
            problem_faced_by=request.form.get("problem_faced_by", ""),
            app_user_information=request.form.get("app_user_information", ""),
            app_version=request.form.get("app_version", ""),
            product=request.form.get("product", ""),
            dso_name=request.form.get("dso_name", ""),
            form_submit_email=request.form.get(
                "form_submit_email", current_user.email or ""
            ),
            priority=request.form.get("priority", "Medium"),
            comments=request.form.get("comments", ""),
            screenshot_path=screenshot_path,
        )
        db.session.add(ticket)
        db.session.flush()  # get ticket.id before commit

        # Apply SLA due date
        sla = SLAPolicy.query.filter_by(priority=ticket.priority, is_active=True).first()
        if sla:
            ticket.due_date = datetime.utcnow() + timedelta(hours=sla.resolution_hours)

        # Auto-assign: pick active DSO/Admin with fewest open tickets (round-robin)
        if not ticket.assigned_to_id:
            agents = User.query.filter(
                User.role.in_([Role.DSO, Role.ADMIN]), User.is_active == True
            ).all()
            if agents:
                best = min(
                    agents,
                    key=lambda u: Ticket.query.filter(
                        Ticket.assigned_to_id == u.id,
                        Ticket.current_status.notin_(["Resolved", "Closed"])
                    ).count()
                )
                ticket.assigned_to_id = best.id

        # Handle additional file attachments
        upload_folder = current_app.config["UPLOAD_FOLDER"]
        os.makedirs(upload_folder, exist_ok=True)
        for file in request.files.getlist("attachments"):
            if file and file.filename and allowed_file(file.filename):
                safe, reason = scan_file_magic(file)
                if not safe:
                    flash(f"Attachment '{file.filename}' rejected: {reason}", "warning")
                    continue
                fname = secure_filename(
                    f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                )
                file.save(os.path.join(upload_folder, fname))
                db.session.add(TicketAttachment(
                    ticket_id=ticket.id,
                    filename=f"screenshots/{fname}",
                    original_name=file.filename,
                    file_size=os.path.getsize(os.path.join(upload_folder, fname)),
                    uploaded_by_id=current_user.id,
                ))

        # Save custom field values
        for field in CustomField.query.filter_by(is_active=True).all():
            val = request.form.get(f"cf_{field.id}", "").strip()
            if val:
                db.session.add(TicketFieldValue(ticket_id=ticket.id, field_id=field.id, value=val))

        # Create CSAT placeholder (to be filled when resolved)
        db.session.add(CSATRating(ticket_id=ticket.id))

        log_history(ticket.id, current_user.id, "Ticket created")
        db.session.commit()  # Save ticket immediately before notifications

        # Notifications and automation — wrapped so failures never block ticket creation
        try:
            notify_staff(
                f"New ticket {ticket.sl_no}: {ticket.issue_type} — {ticket.district}",
                ticket_id=ticket.id,
                notif_type="info",
                exclude_user_id=current_user.id,
            )
            if ticket.priority in ("Critical", "Urgent"):
                notify_staff(
                    f"{ticket.priority.upper()} ticket {ticket.sl_no}: {(ticket.problem_details or '')[:80]}",
                    ticket_id=ticket.id,
                    notif_type="danger",
                    roles=[Role.SUPER_ADMIN, Role.ADMIN],
                )
                send_notification(
                    current_app.config["ADMIN_EMAIL"],
                    f"[{ticket.priority.upper()}] New Ticket {ticket.sl_no}",
                    f"A {ticket.priority.lower()} ticket has been raised.\n\n"
                    f"District: {ticket.district}\nIssue: {ticket.problem_details}\n"
                    f"Reporter: {ticket.issue_reporter_name}",
                )
                notify_slack(
                    f":rotating_light: *{ticket.priority} ticket* {ticket.sl_no}: {(ticket.problem_details or '')[:120]}",
                    ticket=ticket, level="danger",
                )
                notify_teams(
                    f"🚨 {ticket.priority} ticket {ticket.sl_no}: {(ticket.problem_details or '')[:120]}",
                    ticket=ticket, level="danger",
                )
            from utils import run_automation_rules
            run_automation_rules(ticket, "ticket_created", current_user.id)
        except Exception as _notify_err:
            current_app.logger.warning(f"Post-creation notification error (ticket saved): {_notify_err}")

        flash(f"Ticket {ticket.sl_no} submitted successfully! Tracking your issue.", "success")
        return redirect(url_for("tickets.detail", ticket_id=ticket.id))

    dso_users = User.query.filter_by(role=Role.DSO, is_active=True).all()
    all_tags = Tag.query.order_by(Tag.name).all()
    from models import IssueCategory, Country
    categories = IssueCategory.query.filter_by(is_active=True).order_by(IssueCategory.display_order).all()
    countries = Country.query.filter_by(is_active=True).order_by(Country.name).all()
    custom_fields = CustomField.query.filter_by(is_active=True).order_by(CustomField.display_order).all()

    user_country_id = getattr(current_user, 'country_id', None)
    user_admin1_id = getattr(current_user, 'admin1_id', None)
    user_admin2_id = getattr(current_user, 'admin2_id', None)
    user_admin3_id = getattr(current_user, 'admin3_id', None)

    return render_template(
        "create_ticket.html",
        dso_users=dso_users,
        statuses=STATUSES,
        priorities=PRIORITIES,
        platforms=PLATFORMS,
        issue_types=ISSUE_TYPES,
        products=PRODUCTS,
        all_tags=all_tags,
        categories=categories,
        countries=countries,
        custom_fields=custom_fields,
        user_country_id=user_country_id,
        user_admin1_id=user_admin1_id,
        user_admin2_id=user_admin2_id,
        user_admin3_id=user_admin3_id,
    )


# ── Detail ─────────────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>")
@login_required
def detail(ticket_id):
    from models import Role
    if current_user.role == Role.VIEWER:
        flash("Viewers do not have access to individual ticket details.", "warning")
        return redirect(url_for("dashboard.main"))
    ticket = Ticket.query.get_or_404(ticket_id)
    assignable = User.query.filter(
        User.role.in_([Role.DSO, Role.ADMIN, Role.SUPER_ADMIN]), User.is_active == True
    ).all()
    can_see_internal = current_user.can_update_tickets()
    comments = [
        c for c in ticket.ticket_comments
        if not c.is_internal or can_see_internal
    ]
    is_watching = TicketWatcher.query.filter_by(
        ticket_id=ticket_id, user_id=current_user.id
    ).first() is not None
    all_tags = Tag.query.order_by(Tag.name).all()
    ticket_tag_ids = {t.id for t in ticket.tags}
    canned = (
        CannedResponse.query.filter_by(is_active=True)
        .order_by(CannedResponse.category, CannedResponse.title).all()
        if can_see_internal else []
    )
    custom_fields = CustomField.query.filter_by(is_active=True).order_by(CustomField.display_order).all()
    field_values = {fv.field_id: fv.value for fv in ticket.custom_values}
    outgoing = TicketLink.query.filter_by(source_id=ticket_id).all()
    incoming = TicketLink.query.filter_by(target_id=ticket_id).all()
    time_total = sum(e.minutes for e in ticket.time_entries)
    call_logs = CallLog.query.filter_by(ticket_id=ticket_id).order_by(CallLog.logged_at.desc()).all()

    # Collision: acquire lock or detect existing lock
    lock_warning = None
    LOCK_TTL = 5 * 60  # 5 minutes
    now = datetime.utcnow()
    if ticket.locked_by_id and ticket.locked_by_id != current_user.id:
        if ticket.locked_at and (now - ticket.locked_at).total_seconds() < LOCK_TTL:
            lock_warning = ticket.locked_by
        else:
            ticket.locked_by_id = None
            ticket.locked_at = None
            db.session.commit()
    if not ticket.locked_by_id or ticket.locked_by_id == current_user.id:
        ticket.locked_by_id = current_user.id
        ticket.locked_at = now
        db.session.commit()

    return render_template(
        "ticket_detail.html",
        ticket=ticket,
        comments=comments,
        history=ticket.ticket_history,
        assignable=assignable,
        statuses=STATUSES,
        priorities=PRIORITIES,
        can_update=current_user.can_update_tickets(),
        is_admin=current_user.is_admin(),
        can_see_internal=can_see_internal,
        is_watching=is_watching,
        all_tags=all_tags,
        ticket_tag_ids=ticket_tag_ids,
        canned_responses=canned,
        custom_fields=custom_fields,
        field_values=field_values,
        outgoing_links=outgoing,
        incoming_links=incoming,
        time_entries=ticket.time_entries,
        time_total=time_total,
        call_logs=call_logs,
        lock_warning=lock_warning,
    )


# ── Update ─────────────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/update", methods=["POST"])
@login_required
def update(ticket_id):
    if not current_user.can_update_tickets():
        flash("You don't have permission to update tickets.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))

    ticket = Ticket.query.get_or_404(ticket_id)

    old_status    = ticket.current_status
    old_priority  = ticket.priority
    old_assigned  = ticket.assigned_to_id
    old_escalation = ticket.escalation_level

    # Apply changes
    ticket.current_status  = request.form.get("current_status", ticket.current_status)
    ticket.priority        = request.form.get("priority", ticket.priority)
    ticket.remarks         = request.form.get("remarks", "")
    ticket.dso_name        = request.form.get("dso_name", ticket.dso_name)
    ticket.escalation_level = int(request.form.get("escalation_level", ticket.escalation_level))
    ticket.updated_at      = datetime.utcnow()

    assigned_raw = request.form.get("assigned_to_id", "")
    ticket.assigned_to_id = int(assigned_raw) if assigned_raw and assigned_raw != "0" else None

    if request.form.get("solved") == "yes" and not ticket.solved_status:
        ticket.solved_status  = True
        ticket.current_status = "Resolved"
        ticket.solved_date    = datetime.utcnow()
        ticket.solved_by      = current_user.full_name

    # ── History & notifications ──
    if old_status != ticket.current_status:
        log_history(ticket.id, current_user.id, "Status changed", old_status, ticket.current_status)
        if ticket.reporter_id and ticket.reporter_id != current_user.id:
            create_notification(
                ticket.reporter_id,
                f"Ticket {ticket.sl_no}: status changed from {old_status} to {ticket.current_status}",
                ticket_id=ticket.id,
                notif_type="info",
            )
        if ticket.form_submit_email:
            send_ticket_notification(
                ticket.form_submit_email,
                ticket.issue_reporter_name or "User",
                ticket.sl_no, old_status, ticket.current_status, ticket.remarks,
            )

    if old_priority != ticket.priority:
        log_history(ticket.id, current_user.id, "Priority changed", old_priority, ticket.priority)
        notify_staff(
            f"Ticket {ticket.sl_no} priority changed: {old_priority} → {ticket.priority}",
            ticket_id=ticket.id,
            notif_type="warning" if ticket.priority in ("High", "Critical") else "info",
            exclude_user_id=current_user.id,
        )

    if old_assigned != ticket.assigned_to_id and ticket.assigned_to_id:
        log_history(ticket.id, current_user.id, "Assigned", None, str(ticket.assigned_to_id))
        create_notification(
            ticket.assigned_to_id,
            f"Ticket {ticket.sl_no} ({ticket.issue_type}) has been assigned to you",
            ticket_id=ticket.id,
            notif_type="warning",
        )

    if ticket.escalation_level > old_escalation:
        log_history(
            ticket.id, current_user.id,
            f"Escalated to Level {ticket.escalation_level}",
            str(old_escalation), str(ticket.escalation_level),
        )
        notify_staff(
            f"Ticket {ticket.sl_no} escalated to Level {ticket.escalation_level} by {current_user.full_name}",
            ticket_id=ticket.id,
            notif_type="danger",
            roles=[Role.SUPER_ADMIN],
        )

    # Track first response time
    if not ticket.first_response_at and current_user.role != Role.REPORTER:
        ticket.first_response_at = datetime.utcnow()

    # Check SLA breach
    if ticket.due_date and datetime.utcnow() > ticket.due_date and not ticket.sla_breached:
        ticket.sla_breached = True
        notify_staff(
            f"SLA BREACHED: Ticket {ticket.sl_no} is overdue!",
            ticket_id=ticket.id,
            notif_type="danger",
            roles=[Role.SUPER_ADMIN, Role.ADMIN],
        )
        notify_slack(f":alarm_clock: SLA BREACHED: {ticket.sl_no}", ticket=ticket, level="danger")
        notify_teams(f"⏰ SLA BREACHED: {ticket.sl_no}", ticket=ticket, level="danger")

    # Notify watchers on any change
    for watcher in ticket.watchers:
        if watcher.user_id != current_user.id:
            create_notification(
                watcher.user_id,
                f"Ticket {ticket.sl_no} was updated by {current_user.full_name}",
                ticket_id=ticket.id,
                notif_type="info",
            )

    if ticket.current_status == "Resolved":
        log_history(ticket.id, current_user.id, "Ticket resolved", None, current_user.full_name)
        # Send CSAT rating request
        csat = ticket.csat_rating
        if not csat:
            from models import CSATRating
            csat = CSATRating(ticket_id=ticket.id)
            db.session.add(csat)
            db.session.flush()
        if ticket.form_submit_email and not csat.submitted_at:
            rating_url = f"{request.host_url.rstrip('/')}/admin/rate/{csat.token}"
            _send_csat_email(ticket, csat.token, rating_url)
        # WhatsApp CSAT nudge — send if contact is a phone number (not email)
        contact = ticket.issue_reporter_contact
        if contact and not "@" in contact:
            _send_wa_csat(ticket, csat.token)

    from utils import run_automation_rules
    if old_status != ticket.current_status:
        run_automation_rules(ticket, "status_changed", current_user.id)
    if old_priority != ticket.priority:
        run_automation_rules(ticket, "priority_changed", current_user.id)
    db.session.commit()
    flash("Ticket updated successfully!", "success")
    return redirect(url_for("tickets.detail", ticket_id=ticket_id))


def _send_csat_email(ticket, token, rating_url):
    html = f"""
<div style="font-family:Arial,sans-serif;max-width:500px;margin:auto;padding:24px;">
  <h2 style="color:#1d6fa4;">How did we do?</h2>
  <p>Your ticket <strong>{ticket.sl_no}</strong> has been resolved.</p>
  <p>Please take a moment to rate your support experience:</p>
  <div style="text-align:center;margin:24px 0;">
    {''.join(f'<a href="{rating_url}?r={i}" style="display:inline-block;margin:4px;padding:10px 18px;background:#1d6fa4;color:#fff;border-radius:6px;text-decoration:none;font-size:1.2rem;">{"⭐"*i}</a>' for i in range(1, 6))}
  </div>
  <p style="color:#6b7280;font-size:.85rem;">Or <a href="{rating_url}">click here</a> to leave detailed feedback.</p>
</div>"""
    send_notification(
        ticket.form_submit_email,
        f"How did we do? Rate your experience — {ticket.sl_no}",
        f"Your ticket {ticket.sl_no} has been resolved. Please rate us at: {rating_url}",
        html_body=html,
    )


def _send_wa_csat(ticket, token):
    try:
        from routes.webhooks import _wa_send
        phone = ticket.issue_reporter_contact
        if not phone:
            return
        # Update WhatsApp session to CSAT_PENDING state
        from models import WhatsAppSession
        session = WhatsAppSession.query.filter_by(phone=phone).first()
        if session:
            session.state = "CSAT_PENDING"
            session.data = {**(session.data or {}), "ticket_id": ticket.id}
        msg = (
            f"Hi! Your ticket *{ticket.sl_no}* has been resolved by our support team.\n\n"
            f"How satisfied are you with the support you received?\n"
            f"Reply with a number:\n"
            f"5⭐ Excellent\n4⭐ Good\n3⭐ Average\n2⭐ Poor\n1⭐ Very Poor"
        )
        _wa_send(phone, msg)
        # Log the nudge
        from models import NudgeLog
        db.session.add(NudgeLog(
            ticket_id=ticket.id,
            nudge_type="csat_whatsapp",
            recipient=phone,
            channel="whatsapp",
            message_preview=f"CSAT survey for {ticket.sl_no}",
            status="sent"
        ))
    except Exception as e:
        current_app.logger.warning(f"WA CSAT send failed: {e}")


# ── Comment ────────────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/comment", methods=["POST"])
@login_required
def add_comment(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    body = request.form.get("body", "").strip()
    if not body:
        flash("Comment cannot be empty.", "warning")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))

    is_internal = (
        request.form.get("is_internal") == "on"
        and current_user.can_update_tickets()
    )
    comment = TicketComment(
        ticket_id=ticket.id,
        author_id=current_user.id,
        body=body,
        is_internal=is_internal,
    )
    db.session.add(comment)
    db.session.flush()

    # Handle file attachment on comment
    file = request.files.get("attachment")
    if file and file.filename and allowed_file(file.filename):
        safe, reason = scan_file_magic(file)
        if not safe:
            flash(f"Attachment rejected: {reason}", "warning")
        else:
            upload_folder = current_app.config["UPLOAD_FOLDER"]
            os.makedirs(upload_folder, exist_ok=True)
            fname = secure_filename(f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}")
            file.save(os.path.join(upload_folder, fname))
            db.session.add(TicketAttachment(
                ticket_id=ticket.id,
                comment_id=comment.id,
                filename=f"screenshots/{fname}",
                original_name=file.filename,
                file_size=os.path.getsize(os.path.join(upload_folder, fname)),
                uploaded_by_id=current_user.id,
            ))

    # Track first response by staff
    if not ticket.first_response_at and current_user.role != Role.REPORTER:
        ticket.first_response_at = datetime.utcnow()

    log_history(
        ticket.id, current_user.id,
        "Internal note added" if is_internal else "Comment added",
    )

    if not is_internal:
        # Notify reporter when staff comments
        if ticket.reporter_id and ticket.reporter_id != current_user.id:
            create_notification(
                ticket.reporter_id,
                f"New reply on ticket {ticket.sl_no} from {current_user.full_name}",
                ticket_id=ticket.id, notif_type="info",
            )
        # Notify staff when reporter comments
        if current_user.role == Role.REPORTER:
            notify_staff(
                f"{current_user.full_name} commented on ticket {ticket.sl_no}",
                ticket_id=ticket.id, notif_type="info",
                exclude_user_id=current_user.id,
            )

    from utils import run_automation_rules
    run_automation_rules(ticket, "reply_received", current_user.id)
    db.session.commit()
    flash("Comment added.", "success")
    return redirect(url_for("tickets.detail", ticket_id=ticket_id) + "#comments")


# ── Delete ─────────────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/delete", methods=["POST"])
@login_required
def delete(ticket_id):
    if current_user.role != Role.SUPER_ADMIN:
        flash("Only Super Admin can delete tickets.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    ticket = Ticket.query.get_or_404(ticket_id)
    db.session.delete(ticket)
    db.session.commit()
    flash("Ticket deleted.", "success")
    return redirect(url_for("dashboard.main"))


# ── Export CSV ─────────────────────────────────────────────────────────────────

@tickets.route("/export")
@login_required
def export_csv():
    if not current_user.can_update_tickets():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard.main"))

    all_tickets = _filtered_query().all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "SL No", "District", "Upazila", "Reporter", "Role", "Contact",
        "Reporting Date", "Issue Start Date", "Platform", "Issue Type",
        "Problem Details", "Problem Faced By", "App Version", "Product",
        "DSO Name", "Status", "Priority", "Escalation Level",
        "Solved", "Solved Date", "Solved By", "Remarks",
    ])
    for t in all_tickets:
        writer.writerow([
            t.sl_no, t.district, t.upazila, t.issue_reporter_name,
            t.issue_reported_by_role, t.issue_reporter_contact,
            t.reporting_date.strftime("%Y-%m-%d %H:%M") if t.reporting_date else "",
            t.issue_start_date.strftime("%Y-%m-%d") if t.issue_start_date else "",
            t.spice_platform, t.issue_type, t.problem_details, t.problem_faced_by,
            t.app_version, t.product, t.dso_name, t.current_status, t.priority,
            t.escalation_level, "Yes" if t.solved_status else "No",
            t.solved_date.strftime("%Y-%m-%d") if t.solved_date else "",
            t.solved_by, t.remarks,
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"tickets_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv",
    )


@tickets.route("/export-excel")
@login_required
def export_excel():
    if not current_user.can_update_tickets():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard.main"))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        flash("Excel export requires openpyxl. Run: pip install openpyxl", "warning")
        return redirect(url_for("dashboard.main"))

    all_tickets = _filtered_query().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    headers = [
        "SL No", "Reporter", "Role", "Contact", "Email",
        "Country", "District", "Upazila",
        "Category", "Subcategory", "Issue Type", "Platform", "Product",
        "Priority", "Status", "Escalation Level",
        "Assigned To", "Reporting Date", "Issue Start Date",
        "First Response", "Due Date", "SLA Breached",
        "Solved", "Solved Date", "Solved By", "Remarks",
        "Problem Details",
    ]

    header_fill = PatternFill(fgColor="1D6FA4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(h) + 2)

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    priority_colors = {
        "Critical": "DC3545", "Urgent": "DC3545",
        "High": "FD7E14", "Medium": "17A2B8", "Low": "28A745",
    }
    status_colors = {
        "Open": "DC3545", "In Progress": "FFC107",
        "Resolved": "28A745", "Closed": "6C757D",
    }

    for row_idx, t in enumerate(all_tickets, 2):
        row_data = [
            t.sl_no, t.issue_reporter_name, t.issue_reported_by_role,
            t.issue_reporter_contact, t.form_submit_email,
            t.ticket_country.name if t.ticket_country else t.district or "",
            t.district, t.upazila,
            t.ticket_category.name if t.ticket_category else "",
            t.ticket_subcategory.name if t.ticket_subcategory else "",
            t.issue_type, t.spice_platform, t.product,
            t.priority, t.current_status, t.escalation_level,
            t.assignee.full_name if t.assignee else "",
            t.reporting_date.strftime("%Y-%m-%d %H:%M") if t.reporting_date else "",
            t.issue_start_date.strftime("%Y-%m-%d") if t.issue_start_date else "",
            t.first_response_at.strftime("%Y-%m-%d %H:%M") if t.first_response_at else "",
            t.due_date.strftime("%Y-%m-%d %H:%M") if t.due_date else "",
            "Yes" if t.sla_breached else "No",
            "Yes" if t.solved_status else "No",
            t.solved_date.strftime("%Y-%m-%d %H:%M") if t.solved_date else "",
            t.solved_by or "",
            t.remarks or "",
            (t.problem_details or "")[:500],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=False)

        # Color-code priority column (col 14)
        p_color = priority_colors.get(t.priority)
        if p_color:
            ws.cell(row=row_idx, column=14).font = Font(bold=True, color=p_color)

        # Zebra striping
        if row_idx % 2 == 0:
            zebra = PatternFill(fgColor="F8FAFC", fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = zebra

    # Problem details column wider
    ws.column_dimensions[get_column_letter(len(headers))].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"tickets_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _filtered_query():
    # Apply regional scope so exports only contain the user's visible tickets
    from routes.dashboard import _region_filter
    q, _ = _region_filter(Ticket.query, current_user)
    if request.args.get("status"):
        q = q.filter(Ticket.current_status == request.args["status"])
    if request.args.get("priority"):
        q = q.filter(Ticket.priority == request.args["priority"])
    if request.args.get("district"):
        q = q.filter(Ticket.district.ilike(f"%{request.args['district']}%"))
    if request.args.get("date_from"):
        try:
            q = q.filter(Ticket.created_at >= datetime.strptime(request.args["date_from"], "%Y-%m-%d"))
        except ValueError:
            pass
    if request.args.get("date_to"):
        try:
            q = q.filter(Ticket.created_at <= datetime.strptime(request.args["date_to"], "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    sla_f = request.args.get("sla")
    if sla_f == "breached":
        q = q.filter(Ticket.sla_breached == True, Ticket.current_status.notin_(["Resolved", "Closed"]))
    elif sla_f == "warning":
        cutoff = datetime.utcnow() + timedelta(hours=4)
        q = q.filter(
            Ticket.due_date.isnot(None),
            Ticket.due_date > datetime.utcnow(),
            Ticket.due_date <= cutoff,
            Ticket.current_status.notin_(["Resolved", "Closed"]),
        )
    if request.args.get("search"):
        term = f"%{request.args['search']}%"
        q = q.filter(db.or_(
            Ticket.problem_details.ilike(term),
            Ticket.issue_reporter_name.ilike(term),
            Ticket.sl_no.ilike(term),
            Ticket.district.ilike(term),
            Ticket.issue_type.ilike(term),
        ))
    return q.order_by(Ticket.created_at.desc())


# ── Call logging ───────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/call", methods=["POST"])
@login_required
def log_call(ticket_id):
    if not current_user.can_update_tickets():
        flash("Access denied.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    Ticket.query.get_or_404(ticket_id)
    caller_name = request.form.get("caller_name", "").strip()
    caller_phone = request.form.get("caller_phone", "").strip()
    direction = request.form.get("direction", "inbound")
    duration = request.form.get("duration_minutes", 0, type=int)
    outcome = request.form.get("outcome", "resolved")
    notes = request.form.get("notes", "").strip()
    db.session.add(CallLog(
        ticket_id=ticket_id,
        agent_id=current_user.id,
        caller_name=caller_name or None,
        caller_phone=caller_phone or None,
        direction=direction,
        duration_minutes=duration,
        outcome=outcome,
        notes=notes or None,
    ))
    log_history(ticket_id, current_user.id, f"Call logged ({direction}, {duration}m, {outcome})")
    db.session.commit()
    flash(f"Call logged ({duration}m).", "success")
    return redirect(url_for("tickets.detail", ticket_id=ticket_id) + "#calls")


@tickets.route("/calls/<int:call_id>/delete", methods=["POST"])
@login_required
def delete_call(call_id):
    call = CallLog.query.get_or_404(call_id)
    tid = call.ticket_id
    if call.agent_id != current_user.id and not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=tid))
    db.session.delete(call)
    db.session.commit()
    flash("Call log deleted.", "success")
    return redirect(url_for("tickets.detail", ticket_id=tid) + "#calls")


# ── Lock release ───────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/unlock", methods=["POST"])
@login_required
def unlock(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.locked_by_id == current_user.id:
        ticket.locked_by_id = None
        ticket.locked_at = None
        db.session.commit()
    return ("", 204)


# ── Duplicate ticket ────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/duplicate", methods=["POST"])
@login_required
def duplicate(ticket_id):
    if not current_user.can_update_tickets():
        flash("Access denied.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    src = Ticket.query.get_or_404(ticket_id)
    new_ticket = Ticket(
        sl_no=Ticket.generate_sl_no(),
        channel=src.channel,
        district=src.district,
        upazila=src.upazila,
        country_id=src.country_id,
        admin1_id=src.admin1_id,
        admin2_id=src.admin2_id,
        admin3_id=src.admin3_id,
        category_id=src.category_id,
        subcategory_id=src.subcategory_id,
        reporter_id=current_user.id,
        issue_reported_by_role=current_user.role.value,
        issue_reporter_name=current_user.full_name,
        issue_start_date=src.issue_start_date,
        spice_platform=src.spice_platform,
        issue_type=src.issue_type,
        problem_details=src.problem_details,
        problem_faced_by=src.problem_faced_by,
        app_user_information=src.app_user_information,
        app_version=src.app_version,
        product=src.product,
        priority=src.priority,
        parent_id=src.id,
    )
    db.session.add(new_ticket)
    db.session.flush()
    sla = SLAPolicy.query.filter_by(priority=new_ticket.priority, is_active=True).first()
    if sla:
        new_ticket.due_date = datetime.utcnow() + timedelta(hours=sla.resolution_hours)
    db.session.add(CSATRating(ticket_id=new_ticket.id))
    log_history(new_ticket.id, current_user.id, f"Duplicated from {src.sl_no}")
    db.session.commit()
    flash(f"Ticket duplicated as {new_ticket.sl_no}.", "success")
    return redirect(url_for("tickets.detail", ticket_id=new_ticket.id))


# ── Merge tickets ───────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/merge", methods=["POST"])
@login_required
def merge(ticket_id):
    if not current_user.can_update_tickets():
        flash("Access denied.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    target_sl = request.form.get("target_sl", "").strip()
    target = Ticket.query.filter_by(sl_no=target_sl).first()
    if not target or target.id == ticket_id:
        flash("Invalid target ticket SL number.", "warning")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    src = Ticket.query.get_or_404(ticket_id)

    # Move comments from src to target
    TicketComment.query.filter_by(ticket_id=src.id).update({"ticket_id": target.id})
    # Move attachments
    TicketAttachment.query.filter_by(ticket_id=src.id).update({"ticket_id": target.id})
    # Close source ticket
    src.current_status = "Closed"
    src.remarks = (src.remarks or "") + f"\n[Merged into {target.sl_no}]"
    src.updated_at = datetime.utcnow()
    log_history(src.id, current_user.id, f"Merged into {target.sl_no}")
    log_history(target.id, current_user.id, f"Received merge from {src.sl_no}")
    db.session.commit()
    flash(f"Ticket {src.sl_no} merged into {target.sl_no}.", "success")
    return redirect(url_for("tickets.detail", ticket_id=target.id))


# ── Ticket links ────────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/link", methods=["POST"])
@login_required
def add_link(ticket_id):
    if not current_user.can_update_tickets():
        flash("Access denied.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    target_sl = request.form.get("link_sl", "").strip()
    link_type = request.form.get("link_type", "related")
    target = Ticket.query.filter_by(sl_no=target_sl).first()
    if not target or target.id == ticket_id:
        flash("Invalid ticket SL number.", "warning")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    existing = TicketLink.query.filter_by(source_id=ticket_id, target_id=target.id).first()
    if not existing:
        db.session.add(TicketLink(
            source_id=ticket_id, target_id=target.id,
            link_type=link_type, created_by_id=current_user.id,
        ))
        db.session.commit()
        flash(f"Linked to {target.sl_no} ({link_type}).", "success")
    else:
        flash("Link already exists.", "warning")
    return redirect(url_for("tickets.detail", ticket_id=ticket_id))


@tickets.route("/links/<int:link_id>/delete", methods=["POST"])
@login_required
def delete_link(link_id):
    if not current_user.can_update_tickets():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard.main"))
    link = TicketLink.query.get_or_404(link_id)
    src_id = link.source_id
    db.session.delete(link)
    db.session.commit()
    flash("Link removed.", "success")
    return redirect(url_for("tickets.detail", ticket_id=src_id))


# ── Time tracking ───────────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/time", methods=["POST"])
@login_required
def log_time(ticket_id):
    if not current_user.can_update_tickets():
        flash("Access denied.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    Ticket.query.get_or_404(ticket_id)
    minutes = request.form.get("minutes", type=int, default=0)
    note = request.form.get("note", "").strip()
    if minutes <= 0:
        flash("Enter a valid number of minutes.", "warning")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    db.session.add(TimeEntry(ticket_id=ticket_id, user_id=current_user.id, minutes=minutes, note=note))
    db.session.commit()
    flash(f"{minutes}m logged.", "success")
    return redirect(url_for("tickets.detail", ticket_id=ticket_id) + "#time")


@tickets.route("/time/<int:entry_id>/delete", methods=["POST"])
@login_required
def delete_time(entry_id):
    entry = TimeEntry.query.get_or_404(entry_id)
    tid = entry.ticket_id
    if entry.user_id != current_user.id and not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("tickets.detail", ticket_id=tid))
    db.session.delete(entry)
    db.session.commit()
    flash("Time entry deleted.", "success")
    return redirect(url_for("tickets.detail", ticket_id=tid) + "#time")


# ── REST API ───────────────────────────────────────────────────────────────────

@tickets.route("/api/create", methods=["POST"])
def api_create():
    api_key = request.headers.get("X-API-Key")
    if not api_key or api_key != current_app.config["API_KEY"]:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json or {}
    for field in ["district", "issue_type", "problem_details"]:
        if not data.get(field):
            return jsonify({"error": f"Missing required field: {field}"}), 400

    ticket = Ticket(
        sl_no=Ticket.generate_sl_no(),
        district=data.get("district"),
        upazila=data.get("upazila"),
        issue_type=data.get("issue_type"),
        problem_details=data.get("problem_details"),
        spice_platform=data.get("spice_platform"),
        priority=data.get("priority", "Medium"),
        issue_reporter_name=data.get("reporter_name", "API"),
        issue_reporter_contact=data.get("contact"),
        form_submit_email=data.get("email"),
        product=data.get("product"),
    )
    db.session.add(ticket)
    db.session.commit()
    return jsonify({"status": "success", "ticket_id": ticket.id, "sl_no": ticket.sl_no}), 201


# ── Ticket preview panel (AJAX) ────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/preview")
@login_required
def preview(ticket_id):
    """Returns a lightweight partial HTML for the split-view preview panel."""
    from models import Role
    ticket = Ticket.query.get_or_404(ticket_id)
    # Reporters can only preview their own tickets
    if current_user.role == Role.REPORTER and ticket.reporter_id != current_user.id:
        return ("<div class='p-4 text-muted'>Access denied.</div>", 403)
    comments = [c for c in ticket.ticket_comments
                if not c.is_internal or current_user.can_update_tickets()]
    return render_template("tickets/_preview.html",
                           ticket=ticket, comments=comments[-5:])


# ── Attachment download ─────────────────────────────────────────────────────────

@tickets.route("/<int:ticket_id>/attachments/<int:attachment_id>/download")
@login_required
def download_attachment(ticket_id, attachment_id):
    from models import TicketAttachment, Role
    ticket = Ticket.query.get_or_404(ticket_id)
    attachment = TicketAttachment.query.filter_by(id=attachment_id, ticket_id=ticket_id).first_or_404()
    if current_user.role == Role.VIEWER:
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard.main"))
    if current_user.role == Role.REPORTER and ticket.reporter_id != current_user.id:
        flash("You can only download attachments from your own tickets.", "danger")
        return redirect(url_for("dashboard.main"))
    import pathlib
    file_path = pathlib.Path(current_app.static_folder) / attachment.filename.lstrip("/").replace("screenshots/", "screenshots/")
    if not file_path.exists():
        flash("File not found on server.", "warning")
        return redirect(url_for("tickets.detail", ticket_id=ticket_id))
    as_inline = attachment.filename.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".pdf"))
    return send_file(
        str(file_path),
        as_attachment=not as_inline,
        download_name=attachment.original_name or file_path.name,
    )
